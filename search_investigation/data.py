from __future__ import print_function

import gzip
import os
import shutil

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from .paths import ARRAY_DERIVED_DIR, DATA_DIR


BREAST_XLSX = DATA_DIR / "2026-01-29_Rehan data_breast.xlsx"
OVARIAN_XLSX = DATA_DIR / "2026-01-29_Rehan data_ovarian.xlsx"
ICOGS_LOOKUP_XLSX = DATA_DIR / "2026-01-30_Rehan_icogs_pheno lookup.xlsx"
ONCO_LOOKUP_XLSX = DATA_DIR / "2026-01-30_Rehan_oncoarray_pheno lookup.xlsx"
SEQ_LOOKUP_XLSX = DATA_DIR / "2026-01-30_Rehan_sequencing_pheno lookup.xlsx"
ICOGS_BIM = DATA_DIR / "search_zuberi_icogs.bim"
ICOGS_FAM = DATA_DIR / "search_zuberi_icogs.fam"
ONCO_BIM = DATA_DIR / "search_zuberi_oncoarray.bim"
ONCO_FAM = DATA_DIR / "search_zuberi_oncoarray.fam"
BRIDGES_GENOTYPES = DATA_DIR / "search_bridges_genotypes.csv.gz"
BRIDGES_MISSENSE = DATA_DIR / "search_bridges_missense.csv"
BRIDGES_TRUNCATING = DATA_DIR / "search_bridges_truncating.csv"


def _read_excel_records(path):
    workbook = load_workbook(str(path), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value) if value is not None else "" for value in next(rows)]
    records = []
    for row in rows:
        record = {}
        for index, header in enumerate(headers):
            record[header] = row[index] if index < len(row) else None
        records.append(record)
    return pd.DataFrame.from_records(records)


def _read_lookup(path):
    frame = _read_excel_records(path)
    columns = list(frame.columns)
    frame = frame.rename(columns={columns[0]: "assay_id", columns[1]: "RZ_ID"})
    frame["assay_id"] = frame["assay_id"].astype(str)
    frame["RZ_ID"] = frame["RZ_ID"].astype(str)
    return frame


def _read_fam_ids(path):
    ids = []
    with open(str(path)) as handle:
        for line in handle:
            line = line.strip()
            if line:
                ids.append(line.split()[0])
    return set(ids)


def _count_lines(path):
    count = 0
    with open(str(path)) as handle:
        for _ in handle:
            count += 1
    return count


def _first_row_column_count(path, gz=False):
    opener = gzip.open if gz else open
    with opener(str(path), "rt", encoding="utf-8", errors="replace") as handle:
        header = handle.readline().rstrip("\n\r")
    return header.count(",") + 1


def load_breast_clinical():
    frame = _read_excel_records(BREAST_XLSX)
    frame["RZ_ID"] = frame["RZ_ID"].astype(str)
    return frame


def load_ovarian_clinical():
    frame = _read_excel_records(OVARIAN_XLSX)
    frame["RZ_ID"] = frame["RZ_ID"].astype(str)
    return frame


def load_lookup_tables():
    return {
        "icogs": _read_lookup(ICOGS_LOOKUP_XLSX),
        "oncoarray": _read_lookup(ONCO_LOOKUP_XLSX),
        "sequencing": _read_lookup(SEQ_LOOKUP_XLSX),
    }


def load_sequence_lookup_map():
    lookup = _read_lookup(SEQ_LOOKUP_XLSX)
    return dict(zip(lookup["assay_id"], lookup["RZ_ID"]))


def modality_dimensions():
    lookup_tables = load_lookup_tables()
    icogs_present = _read_fam_ids(ICOGS_FAM)
    onco_present = _read_fam_ids(ONCO_FAM)
    sequencing_lookup = set(lookup_tables["sequencing"]["RZ_ID"].tolist())
    bridges_lookup_map = load_sequence_lookup_map()

    bridges_genotypes = pd.read_csv(
        str(BRIDGES_GENOTYPES),
        compression="gzip",
        usecols=[0],
    )
    bridges_missense = pd.read_csv(str(BRIDGES_MISSENSE), usecols=[1])
    bridges_truncating = pd.read_csv(str(BRIDGES_TRUNCATING), usecols=[1])

    return [
        {
            "modality": "icogs",
            "samples": len(icogs_present),
            "features": _count_lines(ICOGS_BIM),
            "available_patients": len(
                set(
                    lookup_tables["icogs"].loc[
                        lookup_tables["icogs"]["assay_id"].isin(icogs_present), "RZ_ID"
                    ].tolist()
                )
            ),
        },
        {
            "modality": "oncoarray",
            "samples": len(onco_present),
            "features": _count_lines(ONCO_BIM),
            "available_patients": len(
                set(
                    lookup_tables["oncoarray"].loc[
                        lookup_tables["oncoarray"]["assay_id"].isin(onco_present), "RZ_ID"
                    ].tolist()
                )
            ),
        },
        {
            "modality": "sequencing",
            "samples": lookup_tables["sequencing"].shape[0],
            "features": np.nan,
            "available_patients": len(sequencing_lookup),
        },
        {
            "modality": "bridges_genotypes",
            "samples": bridges_genotypes.shape[0],
            "features": _first_row_column_count(BRIDGES_GENOTYPES, gz=True) - 1,
            "available_patients": len(
                set(
                    bridges_lookup_map.get(value)
                    for value in bridges_genotypes.iloc[:, 0].astype(str).tolist()
                    if bridges_lookup_map.get(value)
                )
            ),
        },
        {
            "modality": "bridges_missense",
            "samples": bridges_missense.shape[0],
            "features": _first_row_column_count(BRIDGES_MISSENSE) - 2,
            "available_patients": len(
                set(
                    bridges_lookup_map.get(value)
                    for value in bridges_missense.iloc[:, 0].astype(str).tolist()
                    if bridges_lookup_map.get(value)
                )
            ),
        },
        {
            "modality": "bridges_truncating",
            "samples": bridges_truncating.shape[0],
            "features": _first_row_column_count(BRIDGES_TRUNCATING) - 2,
            "available_patients": len(
                set(
                    bridges_lookup_map.get(value)
                    for value in bridges_truncating.iloc[:, 0].astype(str).tolist()
                    if bridges_lookup_map.get(value)
                )
            ),
        },
    ]


def available_modalities():
    modalities = ["bridges_truncating", "bridges_missense", "bridges_genotypes"]
    if (ARRAY_DERIVED_DIR / "icogs_matrix.tsv.gz").exists():
        modalities.append("icogs")
    if (ARRAY_DERIVED_DIR / "oncoarray_matrix.tsv.gz").exists():
        modalities.append("oncoarray")
    return modalities


def load_modality_dataframe(modality):
    if modality == "bridges_truncating":
        return _load_bridges_dataframe(BRIDGES_TRUNCATING)
    if modality == "bridges_missense":
        return _load_bridges_dataframe(BRIDGES_MISSENSE)
    if modality == "bridges_genotypes":
        return _load_bridges_dataframe(BRIDGES_GENOTYPES, compression="gzip", assay_col="Bridges_ID")
    if modality in ("icogs", "oncoarray"):
        return _load_prepared_array_dataframe(modality)
    raise KeyError("Unknown modality: {0}".format(modality))


def _load_prepared_array_dataframe(modality):
    path = ARRAY_DERIVED_DIR / "{0}_matrix.tsv.gz".format(modality)
    if not path.exists():
        raise FileNotFoundError(
            "Prepared array matrix missing: {0}. Run scripts/prepare_array_matrix.py first.".format(path)
        )
    frame = pd.read_csv(str(path), sep="\t", compression="gzip")
    frame["RZ_ID"] = frame["RZ_ID"].astype(str)
    frame = frame.drop_duplicates("RZ_ID").set_index("RZ_ID")
    feature_columns = [column for column in frame.columns if column != "RZ_ID"]
    return frame[feature_columns].apply(pd.to_numeric, errors="coerce").astype(np.float32)


def _load_bridges_dataframe(path, compression=None, assay_col="BRIDGES_ID"):
    lookup_map = load_sequence_lookup_map()
    frame = pd.read_csv(str(path), compression=compression, low_memory=False)
    drop_columns = [column for column in frame.columns if not str(column) or str(column).startswith("Unnamed:")]
    if drop_columns:
        frame = frame.drop(columns=drop_columns)
    if assay_col not in frame.columns:
        assay_col = frame.columns[0]
    frame[assay_col] = frame[assay_col].astype(str)
    frame["RZ_ID"] = frame[assay_col].map(lookup_map)
    frame = frame[frame["RZ_ID"].notnull()].copy()
    feature_columns = [column for column in frame.columns if column not in (assay_col, "RZ_ID")]
    numeric = frame[feature_columns].apply(pd.to_numeric, errors="coerce")
    numeric["RZ_ID"] = frame["RZ_ID"].astype(str)
    numeric = numeric.groupby("RZ_ID", as_index=False).max()
    numeric = numeric.set_index("RZ_ID")
    return numeric.astype(np.float32)


def prepare_breast_dataset(modality, task_name, clinical_frame, task_transform):
    feature_frame = load_modality_dataframe(modality)
    clinical_frame = clinical_frame.copy()
    clinical_frame["task_label"] = clinical_frame.apply(task_transform, axis=1)
    clinical_frame = clinical_frame[clinical_frame["task_label"].notnull()].copy()
    clinical_frame["RZ_ID"] = clinical_frame["RZ_ID"].astype(str)
    merged = clinical_frame[["RZ_ID", "task_label"]].merge(
        feature_frame.reset_index(),
        on="RZ_ID",
        how="inner",
    )
    if merged.empty:
        raise ValueError(
            "No overlapping patients found for task {0} and modality {1}".format(task_name, modality)
        )
    y = merged["task_label"].copy()
    x = merged.drop(columns=["RZ_ID", "task_label"])
    return x, y


def find_plink_binary():
    for candidate in ("plink2", "plink"):
        found = shutil.which(candidate)
        if found:
            return found
    return None
